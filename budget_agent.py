#!/usr/bin/env python3
"""
Budget Agent
============
Watches Statements Inbox/Inbox/ for new PDF bank statements.
Each new PDF is:
  1. Parsed by Claude AI (extracts all transactions + account info)
  2. Saved as JSON to Statements Inbox/Processed/
  3. Rebuilds Budget Tracker.xlsx from all historical JSONs
  4. Creates/updates Linear issues in the "Budget App" project
  5. Moved to Processed/

Run modes:
  python budget_agent.py            # process any PDFs in Inbox now, then exit
  python budget_agent.py --watch    # stay running, process new PDFs as they arrive

Requirements:
  pip install anthropic pdfplumber openpyxl watchdog
  export ANTHROPIC_API_KEY="sk-ant-..."
  export ZAPIER_API_KEY="..."   # optional — creates Linear issues via Zapier MCP
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import date
from pathlib import Path

# Zapier → Linear integration (optional — only active when ZAPIER_WEBHOOK_URL is set)
try:
    from zapier_client import post_statement_issues as _zapier_post
    _ZAPIER_AVAILABLE = True
except ImportError:
    _ZAPIER_AVAILABLE = False

# ── Paths ────────────────────────────────────────────────────────────────────
DRIVE_ROOT = Path("/Users/gustavooviedo/Library/CloudStorage/GoogleDrive-gustavobills7@gmail.com/My Drive")
INBOX      = DRIVE_ROOT / "Statements Inbox" / "Inbox"
PROCESSED  = DRIVE_ROOT / "Statements Inbox" / "Processed"
EXCEL_OUT  = DRIVE_ROOT / "Statements Inbox" / "Budget Tracker.xlsx"

# ── Colours (openpyxl hex, no #) ─────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GREEN = "E2EFDA"
RED         = "C00000"
LIGHT_RED   = "FDECEA"
GRAY        = "F2F2F2"
WHITE       = "FFFFFF"
ORANGE      = "C55A11"

# ══════════════════════════════════════════════════════════════════════════════
# PDF → Claude parsing
# ══════════════════════════════════════════════════════════════════════════════
PARSE_PROMPT = """You are a financial data extractor. Parse this bank/credit card statement and return ONLY valid JSON.

Filename : {filename}
Today    : {today}

Statement text:
{text}

Return this exact structure (no markdown fences, no extra text):
{{
  "account_name"          : "e.g. Capital One Platinum",
  "account_last4"         : "4-digit string",
  "holder"                : "full name on card",
  "bank"                  : "issuing bank name",
  "card_type"             : "e.g. World Mastercard",
  "statement_period_start": "YYYY-MM-DD",
  "statement_period_end"  : "YYYY-MM-DD",
  "due_date"              : "e.g. May 08, 2026",
  "previous_balance"      : 0.00,
  "payments_total"        : 0.00,
  "purchases_total"       : 0.00,
  "fees_total"            : 0.00,
  "interest_total"        : 0.00,
  "new_balance"           : 0.00,
  "min_payment"           : 0.00,
  "past_due"              : 0.00,
  "credit_limit"          : 0.00,
  "available_credit"      : 0.00,
  "apr"                   : 0.00,
  "overlimit"             : 0.00,
  "transactions": [
    {{
      "transaction_date": "YYYY-MM-DD",
      "post_date"       : "YYYY-MM-DD",
      "description"     : "merchant or charge description",
      "amount"          : 12.34,
      "type"            : "purchase|payment|fee|interest|credit",
      "category"        : "Food & Dining|Transport|Shopping|Bills & Utilities|Entertainment|Travel|Health|Church & Donations|Bank Fees|Interest|Payment|Other"
    }}
  ]
}}

Rules:
- Purchases, fees, interest → positive amounts
- Payments, credits → negative amounts
- Include ALL line items (fees, interest, purchases, payments)
- Infer the full 4-digit year from context in the statement
- Use null for any field you cannot find
"""


def extract_pdf_text(pdf_path: Path) -> str:
    import pdfplumber
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
    return "\n\n--- PAGE BREAK ---\n\n".join(pages)


def parse_with_claude(text: str, filename: str) -> dict:
    import anthropic
    client = anthropic.Anthropic()
    resp = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=4096,
        messages=[{"role": "user", "content": PARSE_PROMPT.format(
            filename=filename,
            today=date.today().isoformat(),
            text=text,
        )}],
    )
    raw = resp.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


# ══════════════════════════════════════════════════════════════════════════════
# Excel generation (rebuilds from all JSONs in Processed/)
# ══════════════════════════════════════════════════════════════════════════════
def load_all_statements() -> list[dict]:
    """Load every JSON in Processed/, sorted by statement end date."""
    stmts = []
    for jf in sorted(PROCESSED.glob("*.json")):
        try:
            with open(jf) as f:
                stmts.append(json.load(f))
        except Exception as e:
            print(f"  Warning: could not load {jf.name}: {e}")
    return stmts


def build_excel(stmts: list[dict]):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(c):  return PatternFill("solid", fgColor=c)
    def fnt(bold=False, color="000000", size=11, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
    def bdr():
        s = Side(border_style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)
    def right(ws, cell):
        ws[cell].alignment = Alignment(horizontal="right", vertical="center")
    def center_cell(ws, cell):
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
    def money_fmt(cell): cell.number_format = '"$"#,##0.00'

    # Group by account (last4) – keep only the most-recent statement per account
    by_account: dict[str, dict] = {}
    for s in stmts:
        key = s.get("account_last4", "????")
        prev = by_account.get(key)
        if prev is None or (s.get("statement_period_end","") > prev.get("statement_period_end","")):
            by_account[key] = s
    accounts = list(by_account.values())

    wb = Workbook()
    wb.remove(wb.active)

    # ── TAB COLOURS per account ──────────────────────────────────────────────
    tab_palette = ["2E75B6", "C55A11", "375623", "7030A0", "843C0C", "1F3864"]

    # ── SUMMARY sheet ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = DARK_BLUE
    ws.sheet_view.showGridLines = False

    col_w = [22, 28, 10, 16, 16, 14, 10, 14, 14, 14]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # title
    ws.merge_cells("A1:J1")
    ws["A1"] = "BUDGET TRACKER — ACCOUNT SUMMARY"
    ws["A1"].font      = fnt(bold=True, color=WHITE, size=16)
    ws["A1"].fill      = fill(DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generated {date.today().strftime('%B %d, %Y')}"
    ws["A2"].font      = fnt(italic=True, color=WHITE, size=11)
    ws["A2"].fill      = fill(MID_BLUE)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    # headers
    hdrs = ["Account Holder","Account / Card","Last 4",
            "Balance ($)","Min. Payment ($)","Due Date",
            "APR (%)","Interest ($)","Past Due ($)","Overlimit ($)"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = fnt(bold=True, color=WHITE, size=10)
        c.fill = fill(MID_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr()
    ws.row_dimensions[4].height = 28

    row_bg = [LIGHT_BLUE, GRAY]
    for idx, acct in enumerate(accounts):
        r  = 5 + idx
        bg = row_bg[idx % 2]
        vals = [
            acct.get("holder",""),
            f"{acct.get('account_name','')} ({acct.get('bank','')})",
            f"...{acct.get('account_last4','')}",
            acct.get("new_balance", 0),
            acct.get("min_payment", 0),
            acct.get("due_date", ""),
            acct.get("apr", 0),
            acct.get("interest_total", 0),
            acct.get("past_due", 0),
            acct.get("overlimit", 0),
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill   = fill(bg)
            c.border = bdr()
            c.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22

        for col in [4, 5, 8, 9, 10]:
            c = ws.cell(row=r, column=col)
            money_fmt(c)
            c.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=7).number_format = '0.00"%"'
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center", vertical="center")

        if (acct.get("past_due") or 0) > 0:
            ws.cell(row=r, column=9).font = fnt(bold=True, color=RED)
        if (acct.get("overlimit") or 0) > 0:
            ws.cell(row=r, column=10).font = fnt(bold=True, color=RED)

    # totals
    tr = 5 + len(accounts)
    ws.row_dimensions[tr].height = 26
    totals = {
        4:  sum(a.get("new_balance",0)     for a in accounts),
        5:  sum(a.get("min_payment",0)     for a in accounts),
        8:  sum(a.get("interest_total",0)  for a in accounts),
        9:  sum(a.get("past_due",0)        for a in accounts),
        10: sum(a.get("overlimit",0)       for a in accounts),
    }
    for col in range(1, 11):
        c = ws.cell(row=tr, column=col)
        c.value = "TOTALS" if col == 1 else totals.get(col, "")
        c.fill   = fill(DARK_BLUE)
        c.font   = fnt(bold=True, color=WHITE, size=11)
        c.border = bdr()
        c.alignment = Alignment(horizontal="right" if col >= 4 else "left", vertical="center")
        if isinstance(c.value, float): money_fmt(c)
    ws.cell(row=tr, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # alerts
    alerts = [(a.get("holder",""), a.get("account_name",""), a.get("account_last4",""),
               a.get("past_due",0), a.get("overlimit",0), a.get("min_payment",0), a.get("due_date",""))
              for a in accounts if (a.get("past_due",0) > 0 or a.get("overlimit",0) > 0)]

    if alerts:
        ar = tr + 2
        ws.merge_cells(f"A{ar}:J{ar}")
        ws[f"A{ar}"] = "⚠  ALERTS"
        ws[f"A{ar}"].font      = fnt(bold=True, color=WHITE, size=11)
        ws[f"A{ar}"].fill      = fill(RED)
        ws[f"A{ar}"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[ar].height = 20

        for i, (holder, aname, last4, past, over, minpay, due) in enumerate(alerts):
            r = ar + 1 + i
            parts = []
            if past  > 0: parts.append(f"PAST DUE ${past:,.2f}")
            if over  > 0: parts.append(f"OVERLIMIT ${over:,.2f}")
            msg = f"{', '.join(parts)}. Min payment ${minpay:,.2f} due {due}."
            ws.merge_cells(f"A{r}:C{r}")
            ws.cell(row=r, column=1, value=f"{holder} — {aname} ...{last4}").font = fnt(bold=True, color=RED)
            ws.merge_cells(f"D{r}:J{r}")
            ws.cell(row=r, column=4, value=msg).font = fnt(color="7F0000")
            for col in range(1, 11):
                ws.cell(row=r, column=col).fill   = fill(LIGHT_RED)
                ws.cell(row=r, column=col).border = bdr()
                ws.cell(row=r, column=col).alignment = Alignment(vertical="center")
            ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A5"

    # ── INDIVIDUAL ACCOUNT sheets ─────────────────────────────────────────────
    for idx, acct in enumerate(accounts):
        tab_color = tab_palette[idx % len(tab_palette)]
        sname = f"...{acct.get('account_last4','')} {acct.get('account_name','').split()[0]}"
        ws2 = wb.create_sheet(sname)
        ws2.sheet_properties.tabColor = tab_color
        ws2.sheet_view.showGridLines = False

        for i, w in enumerate([18, 18, 36, 16, 16, 18], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # title
        ws2.merge_cells("A1:F1")
        ws2["A1"] = f"{acct.get('account_name','')}  —  {acct.get('holder','')}"
        ws2["A1"].font      = fnt(bold=True, color=WHITE, size=14)
        ws2["A1"].fill      = fill(tab_color)
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 32

        ws2.merge_cells("A2:F2")
        period = f"{acct.get('statement_period_start','')} – {acct.get('statement_period_end','')}"
        ws2["A2"] = f"Statement: {period}   |   Card ending ...{acct.get('account_last4','')}"
        ws2["A2"].font      = fnt(italic=True, color=WHITE, size=10)
        ws2["A2"].fill      = fill(tab_color)
        ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[2].height = 18
        ws2.row_dimensions[3].height = 8

        # account info
        info = [
            ("Account Holder",   acct.get("holder","")),
            ("Bank",             acct.get("bank","")),
            ("Card Type",        acct.get("card_type","")),
            ("Statement Period",  period),
            ("Payment Due Date", acct.get("due_date","")),
            ("APR",              acct.get("apr", 0)),
            ("Credit Limit",     acct.get("credit_limit", 0)),
            ("Available Credit", acct.get("available_credit", 0)),
        ]
        for i, (label, val) in enumerate(info):
            r = 4 + i
            ws2.row_dimensions[r].height = 20
            lc = ws2.cell(row=r, column=1, value=label)
            lc.font = fnt(bold=True, color=WHITE, size=10)
            lc.fill = fill(MID_BLUE); lc.border = bdr()
            lc.alignment = Alignment(vertical="center")
            ws2.merge_cells(f"A{r}:B{r}")
            vc = ws2.cell(row=r, column=3, value=val)
            vc.fill = fill(LIGHT_BLUE); vc.border = bdr()
            vc.alignment = Alignment(vertical="center")
            ws2.merge_cells(f"C{r}:D{r}")
            if label == "APR": vc.number_format = '0.00"%"'
            if label in ("Credit Limit", "Available Credit"): money_fmt(vc)

        ws2.row_dimensions[12].height = 10

        # activity summary
        activity = [
            ("Previous Balance", acct.get("previous_balance", 0),   False),
            ("Payments / Credits",acct.get("payments_total",   0),   False),
            ("Purchases",         acct.get("purchases_total",  0),   False),
            ("Fees Charged",      acct.get("fees_total",       0),   False),
            ("Interest Charged",  acct.get("interest_total",   0),   False),
            ("New Balance",       acct.get("new_balance",      0),   True),
        ]
        ws2.merge_cells("A13:B13")
        ws2["A13"] = "ACCOUNT ACTIVITY"
        ws2["A13"].font      = fnt(bold=True, color=WHITE)
        ws2["A13"].fill      = fill(tab_color)
        ws2["A13"].alignment = Alignment(horizontal="center", vertical="center")
        ws2["A13"].border    = bdr()
        ws2.merge_cells("C13:D13")
        ws2["C13"].fill = fill(tab_color); ws2["C13"].border = bdr()
        ws2.row_dimensions[13].height = 22

        for i, (label, val, is_total) in enumerate(activity):
            r = 14 + i
            bg = DARK_BLUE if is_total else (LIGHT_BLUE if i % 2 == 0 else GRAY)
            fc = WHITE if is_total else "000000"
            ws2.row_dimensions[r].height = 20
            lc = ws2.cell(row=r, column=1, value=label)
            lc.font = fnt(bold=is_total, color=fc); lc.fill = fill(bg)
            lc.border = bdr(); lc.alignment = Alignment(vertical="center")
            ws2.merge_cells(f"A{r}:B{r}")
            vc = ws2.cell(row=r, column=3, value=val)
            vc.font = fnt(bold=is_total, color=fc if val >= 0 else ("375623" if not is_total else WHITE))
            vc.fill = fill(bg); vc.border = bdr()
            money_fmt(vc); vc.alignment = Alignment(horizontal="right", vertical="center")
            ws2.merge_cells(f"C{r}:D{r}")

        # payment info
        ps = 14 + len(activity) + 1
        ws2.row_dimensions[ps - 1].height = 10
        payment_rows = [
            ("Minimum Payment Due", acct.get("min_payment", 0),       False),
            ("Past Due Amount",     acct.get("past_due",    0),        acct.get("past_due",0) > 0),
            ("Overlimit Amount",    acct.get("overlimit",   0),        acct.get("overlimit",0) > 0),
        ]
        ws2.merge_cells(f"A{ps}:D{ps}")
        ws2[f"A{ps}"] = "PAYMENT INFORMATION"
        ws2[f"A{ps}"].font      = fnt(bold=True, color=WHITE)
        ws2[f"A{ps}"].fill      = fill(tab_color)
        ws2[f"A{ps}"].alignment = Alignment(horizontal="center", vertical="center")
        ws2[f"A{ps}"].border    = bdr()
        ws2.row_dimensions[ps].height = 22

        for i, (label, val, is_alert) in enumerate(payment_rows):
            r = ps + 1 + i
            bg = LIGHT_RED if is_alert else (LIGHT_BLUE if i % 2 == 0 else GRAY)
            fc = RED if is_alert else "000000"
            ws2.row_dimensions[r].height = 20
            lc = ws2.cell(row=r, column=1, value=label)
            lc.font = fnt(bold=is_alert, color=fc); lc.fill = fill(bg)
            lc.border = bdr(); lc.alignment = Alignment(vertical="center")
            ws2.merge_cells(f"A{r}:B{r}")
            vc = ws2.cell(row=r, column=3, value=val)
            vc.font = fnt(bold=is_alert, color=fc); vc.fill = fill(bg)
            vc.border = bdr(); money_fmt(vc)
            vc.alignment = Alignment(horizontal="right", vertical="center")
            ws2.merge_cells(f"C{r}:D{r}")

        # transactions table
        tx_start = ps + len(payment_rows) + 3
        ws2.row_dimensions[tx_start - 1].height = 10
        for col, h in enumerate(["Trans Date","Post Date","Description","Amount ($)","Type","Category"], 1):
            c = ws2.cell(row=tx_start, column=col, value=h)
            c.font = fnt(bold=True, color=WHITE, size=10)
            c.fill = fill(MID_BLUE); c.border = bdr()
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[tx_start].height = 22

        for i, tx in enumerate(acct.get("transactions", [])):
            r  = tx_start + 1 + i
            amt = tx.get("amount", 0)
            typ = tx.get("type","")
            bg  = LIGHT_GREEN if amt < 0 else (LIGHT_RED if typ in ("fee","interest") else (GRAY if i%2 else WHITE))
            ws2.row_dimensions[r].height = 20
            for col, val in enumerate([
                tx.get("transaction_date",""), tx.get("post_date",""),
                tx.get("description",""), amt, typ, tx.get("category","")
            ], 1):
                c = ws2.cell(row=r, column=col, value=val)
                c.fill = fill(bg); c.border = bdr()
                c.alignment = Alignment(vertical="center")
            ac = ws2.cell(row=r, column=4)
            money_fmt(ac); ac.alignment = Alignment(horizontal="right", vertical="center")
            ac.font = fnt(color="375623" if amt < 0 else (RED if typ in ("fee","interest") else "000000"))

        ws2.freeze_panes = "A4"

    # ── ALL TRANSACTIONS sheet ────────────────────────────────────────────────
    wt = wb.create_sheet("All Transactions")
    wt.sheet_properties.tabColor = "404040"
    wt.sheet_view.showGridLines = False
    for i, w in enumerate([14,14,36,14,12,20,22,10,24], 1):
        wt.column_dimensions[get_column_letter(i)].width = w

    wt.merge_cells("A1:I1")
    wt["A1"] = "ALL TRANSACTIONS"
    wt["A1"].font = fnt(bold=True, color=WHITE, size=14)
    wt["A1"].fill = fill("404040")
    wt["A1"].alignment = Alignment(horizontal="center", vertical="center")
    wt.row_dimensions[1].height = 30
    wt.row_dimensions[2].height = 8

    all_hdrs = ["Trans Date","Post Date","Description","Amount ($)","Type","Category","Account","Last 4","Statement Period"]
    for col, h in enumerate(all_hdrs, 1):
        c = wt.cell(row=3, column=col, value=h)
        c.font = fnt(bold=True, color=WHITE, size=10)
        c.fill = fill(MID_BLUE); c.border = bdr()
        c.alignment = Alignment(horizontal="center", vertical="center")
    wt.row_dimensions[3].height = 24

    all_rows = []
    for s in stmts:
        period = f"{s.get('statement_period_start','')} → {s.get('statement_period_end','')}"
        for tx in s.get("transactions", []):
            all_rows.append((
                tx.get("transaction_date",""),
                tx.get("post_date",""),
                tx.get("description",""),
                tx.get("amount", 0),
                tx.get("type",""),
                tx.get("category",""),
                s.get("account_name",""),
                s.get("account_last4",""),
                period,
            ))
    all_rows.sort(key=lambda x: x[0], reverse=True)

    for i, row_vals in enumerate(all_rows):
        r  = 4 + i
        amt = row_vals[3]
        typ = row_vals[4]
        bg  = LIGHT_GREEN if amt < 0 else (LIGHT_RED if typ in ("fee","interest") else (GRAY if i%2 else WHITE))
        wt.row_dimensions[r].height = 20
        for col, val in enumerate(row_vals, 1):
            c = wt.cell(row=r, column=col, value=val)
            c.fill = fill(bg); c.border = bdr()
            c.alignment = Alignment(vertical="center")
        ac = wt.cell(row=r, column=4)
        money_fmt(ac); ac.alignment = Alignment(horizontal="right", vertical="center")
        ac.font = fnt(color="375623" if amt < 0 else (RED if typ in ("fee","interest") else "000000"))

    wt.freeze_panes = "A4"

    wb.save(EXCEL_OUT)
    print(f"  Excel saved → {EXCEL_OUT}")


# ══════════════════════════════════════════════════════════════════════════════
# Process a single PDF
# ══════════════════════════════════════════════════════════════════════════════
def process_pdf(pdf_path: Path):
    print(f"\n[+] Processing: {pdf_path.name}")
    try:
        print("    Extracting text...")
        text = extract_pdf_text(pdf_path)

        print("    Parsing with Claude AI...")
        data = parse_with_claude(text, pdf_path.name)
        data["source_file"] = pdf_path.name

        print(f"    Account : {data.get('account_name')} ...{data.get('account_last4')}")
        print(f"    Holder  : {data.get('holder')}")
        print(f"    Period  : {data.get('statement_period_start')} → {data.get('statement_period_end')}")
        print(f"    Items   : {len(data.get('transactions', []))}")
        print(f"    Balance : ${data.get('new_balance', 0):,.2f}  |  Min payment: ${data.get('min_payment', 0):,.2f}")

        # Save JSON
        json_path = PROCESSED / (pdf_path.stem + ".json")
        with open(json_path, "w") as f:
            json.dump(data, f, indent=2)
        print(f"    Saved JSON → {json_path.name}")

        # Move PDF to Processed/
        dest = PROCESSED / pdf_path.name
        pdf_path.rename(dest)
        print(f"    Moved PDF → Processed/{pdf_path.name}")

        # Rebuild Excel from all JSONs
        print("    Rebuilding Budget Tracker.xlsx...")
        stmts = load_all_statements()
        build_excel(stmts)

        # Post to Linear via Zapier MCP
        zapier_key = os.environ.get("ZAPIER_API_KEY")
        if _ZAPIER_AVAILABLE and zapier_key:
            print("    Creating Linear issues via Zapier MCP...")
            try:
                _zapier_post(zapier_key, data)
            except Exception as ze:
                print(f"    Zapier warning: {ze}")
        else:
            print("    Zapier: skipped (ZAPIER_API_KEY not set)")

        print(f"    Done ✓")

    except Exception as e:
        print(f"    ERROR: {e}")
        import traceback; traceback.print_exc()


# ══════════════════════════════════════════════════════════════════════════════
# File watcher
# ══════════════════════════════════════════════════════════════════════════════
def run_watch_mode():
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler

    class InboxHandler(FileSystemEventHandler):
        def on_created(self, event):
            if event.is_directory:
                return
            path = Path(event.src_path)
            if path.suffix.lower() == ".pdf":
                # Brief wait to ensure the file is fully written/synced
                time.sleep(2)
                if path.exists():
                    process_pdf(path)

        def on_moved(self, event):
            # Handles files moved/renamed into the folder (e.g. Google Drive sync)
            if event.is_directory:
                return
            path = Path(event.dest_path)
            if path.suffix.lower() == ".pdf":
                time.sleep(2)
                if path.exists():
                    process_pdf(path)

    INBOX.mkdir(parents=True, exist_ok=True)
    observer = Observer()
    observer.schedule(InboxHandler(), str(INBOX), recursive=False)
    observer.start()

    print(f"\nWatching for new PDFs in:\n  {INBOX}")
    print("Drop a bank statement PDF there to process it automatically.")
    print("Press Ctrl+C to stop.\n")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print("\nStopped.")
    observer.join()


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Budget Agent — statement processor")
    parser.add_argument("--watch", action="store_true", help="Keep running and watch Inbox for new PDFs")
    parser.add_argument("--rebuild", action="store_true", help="Rebuild Excel from existing JSONs without processing new PDFs")
    args = parser.parse_args()

    if args.rebuild:
        print("Rebuilding Budget Tracker.xlsx from all processed statements...")
        stmts = load_all_statements()
        if not stmts:
            print("No processed statements found in Processed/.")
            return
        build_excel(stmts)
        print("Done.")
        return

    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: ANTHROPIC_API_KEY not set.\n  export ANTHROPIC_API_KEY='sk-ant-...'")
        sys.exit(1)

    INBOX.mkdir(parents=True, exist_ok=True)
    PROCESSED.mkdir(parents=True, exist_ok=True)

    # Process any PDFs already sitting in Inbox
    pdfs = list(INBOX.glob("*.pdf"))
    if pdfs:
        print(f"Found {len(pdfs)} PDF(s) in Inbox — processing now...")
        for pdf in pdfs:
            process_pdf(pdf)
    else:
        print("Inbox is empty.")

    if args.watch:
        run_watch_mode()
    else:
        if not pdfs:
            print(f"\nTip: run with --watch to stay running and auto-process new PDFs.")


if __name__ == "__main__":
    main()
